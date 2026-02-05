import io
import json
from typing import Dict, Tuple, List

import numpy as np
import pandas as pd
import streamlit as st
import pulp
import openpyxl
from openpyxl.styles import PatternFill, Font


# ----------------------------
# Constants / Month Mapping
# ----------------------------
MONTHS_EN = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
             "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

MONTHS_ES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

ES_TO_IDX = {m: i for i, m in enumerate(MONTHS_ES)}
IDX_TO_EN = {i: m for i, m in enumerate(MONTHS_EN)}
IDX_TO_ES = {i: m for i, m in enumerate(MONTHS_ES)}
EN_TO_IDX = {m: i for i, m in enumerate(MONTHS_EN)}


# ----------------------------
# Helpers
# ----------------------------
def clamp_int(x, default=0, minv=0):
    try:
        xi = int(x)
        return max(minv, xi)
    except Exception:
        return default


def normalize_weights(weights: np.ndarray) -> np.ndarray:
    w = np.array(weights, dtype=float)
    w[w < 0] = 0.0
    s = w.sum()
    if s <= 0:
        return np.ones_like(w) / len(w)
    return w / s


def load_municipios(json_path: str) -> pd.DataFrame:
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    df = pd.DataFrame(data)
    df["codmun"] = df["codmun"].astype(int)
    df["coddep"] = df["coddep"].astype(int)
    df["departamento"] = df["departamento"].astype(str)
    df["municipio_norm"] = df["municipio_norm"].astype(str)
    return df


def excel_month_cols(df: pd.DataFrame) -> Dict[int, str]:
    """
    Detect month column names in the Excel sheet, including 'Julio ' with trailing space.
    Returns dict month_idx -> actual column name.
    """
    colmap = {}
    cols = set(df.columns.astype(str).tolist())
    for i, m in enumerate(MONTHS_ES):
        if m in cols:
            colmap[i] = m
        else:
            for c in cols:
                if str(c).strip() == m:
                    colmap[i] = c
                    break
    return colmap


def _ensure_workbook_in_state(excel_bytes: bytes, filename: str):
    st.session_state.workbook_bytes = excel_bytes
    st.session_state.workbook_filename = filename

def upsert_config_sheet(wb, month_coeffs, cap):
    if "Config" in wb.sheetnames:
        ws = wb["Config"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Config")

    # Header
    ws["A1"] = "key"
    ws["B1"] = "value"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    ws["A2"] = "cap"
    ws["B2"] = int(cap)

    # Efficiency table
    ws["A4"] = "Month"
    ws["B4"] = "Efficiency"
    ws["A4"].font = Font(bold=True)
    ws["B4"].font = Font(bold=True)

    for i in range(12):
        ws.cell(row=5 + i, column=1).value = MONTHS_ES[i]
        ws.cell(row=5 + i, column=2).value = float(month_coeffs[i])


def parse_month_to_idx(x) -> int | None:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    s = str(x).strip()
    if s in EN_TO_IDX:
        return EN_TO_IDX[s]
    # allow Spanish month names in Start/End
    if s in ES_TO_IDX:
        return ES_TO_IDX[s]
    # allow accidental trailing spaces in Spanish
    for k, v in ES_TO_IDX.items():
        if s.strip() == k:
            return v
    return None


def load_state_from_excel_bytes(
    excel_bytes: bytes,
    filename: str,
    mun_catalog: pd.DataFrame,
) -> Dict:
    """
    New format:
      - Plan (or Sheet1): month columns = planned amounts, Start/End define window
      - Tracking: *_delivered and *_finished are the source of truth for delivered/finished
    """
    _ensure_workbook_in_state(excel_bytes, filename)

    xl = pd.ExcelFile(io.BytesIO(excel_bytes))
    sheetnames = xl.sheet_names

    # Read coefficients from Config sheet if present
    if "Config" in xl.sheet_names:
        df_cfg = xl.parse("Config")

        # cap (key/value)
        if "key" in df_cfg.columns and "value" in df_cfg.columns:
            cap_row = df_cfg[df_cfg["key"].astype(str).str.strip().str.lower() == "cap"]
            if len(cap_row) > 0:
                try:
                    st.session_state.capacity = int(cap_row["value"].iloc[0])
                except Exception:
                    pass

        # Month/Efficiency table (Month, Efficiency)
        if "Month" in df_cfg.columns and "Efficiency" in df_cfg.columns:
            coeffs = [None] * 12
            for _, r in df_cfg.iterrows():
                m = str(r.get("Month", "")).strip()
                if m in ES_TO_IDX:
                    idx = ES_TO_IDX[m]
                    try:
                        coeffs[idx] = float(r.get("Efficiency", 0) or 0)
                    except Exception:
                        coeffs[idx] = 0.0
            if all(v is not None for v in coeffs):
                st.session_state.month_coeffs = coeffs

    # Prefer these names
    plan_name = "Plan" if "Plan" in sheetnames else ("Sheet1" if "Sheet1" in sheetnames else sheetnames[0])
    tracking_name = "Tracking" if "Tracking" in sheetnames else None

    df_plan = xl.parse(plan_name).copy()

    if "Codmun" not in df_plan.columns:
        raise ValueError(f"Expected 'Codmun' column in {plan_name}")

    df_plan = df_plan[df_plan["Codmun"].notna()].copy()
    df_plan["Codmun"] = df_plan["Codmun"].astype(int)

    # month column mapping in Plan
    plan_month_colmap = excel_month_cols(df_plan)

    selected_codmun = set(df_plan["Codmun"].tolist())

    # keep only codmun in catalog
    catalog_codmun = set(mun_catalog["codmun"].astype(int).tolist())
    selected_codmun = {c for c in selected_codmun if c in catalog_codmun}

    targets: Dict[int, int] = {}
    windows: Dict[int, Tuple[int, int]] = {}
    frontes: Dict[int, str] = {}

    planned_current = {(c, t): 0 for c in selected_codmun for t in range(12)}
    planned_locked: Dict[Tuple[int, int], int] = {}

    # 1) Read Plan: targets, frente, windows, planned_current
    for _, r in df_plan.iterrows():
        codmun = int(r["Codmun"])
        if codmun not in selected_codmun:
            continue

        # Frente
        frente = r.get("Frente", "")
        frontes[codmun] = "" if pd.isna(frente) else str(frente).strip()

        # Target
        target = r.get("Pisos Proyectados", 0)
        targets[codmun] = clamp_int(target, default=0, minv=0)

        # Start/End (preferred)
        s_idx = parse_month_to_idx(r.get("Start", None))
        e_idx = parse_month_to_idx(r.get("End", None))

        # If missing, infer from non-zero planned months
        if s_idx is None or e_idx is None:
            active = []
            for t in range(12):
                col = plan_month_colmap.get(t)
                if col is None:
                    continue
                v = r.get(col, 0)
                try:
                    vv = int(v or 0)
                except Exception:
                    vv = 0
                if vv > 0:
                    active.append(t)
            if active:
                s_idx, e_idx = min(active), max(active)
            else:
                s_idx, e_idx = 0, 11

        windows[codmun] = (int(s_idx), int(e_idx))

        # planned amounts from Plan month columns
        for t in range(12):
            col = plan_month_colmap.get(t)
            if col is None:
                continue
            v = r.get(col, 0)
            planned_current[(codmun, t)] = clamp_int(v, default=0, minv=0)

    # initialize delivered/finished
    delivered = {(c, t): 0 for c in selected_codmun for t in range(12)}
    finished = {(c, t): False for c in selected_codmun for t in range(12)}

    # 2) Read Tracking (if exists): delivered + finished
    if tracking_name is not None:
        df_track = xl.parse(tracking_name).copy()
        if "Codmun" in df_track.columns:
            df_track = df_track[df_track["Codmun"].notna()].copy()
            df_track["Codmun"] = df_track["Codmun"].astype(int)

            for _, r in df_track.iterrows():
                codmun = int(r["Codmun"])
                if codmun not in selected_codmun:
                    continue

                # Allow Tracking to also set Frente if present (keeps consistency)
                if "Frente" in df_track.columns:
                    frente = r.get("Frente", "")
                    if not pd.isna(frente):
                        frontes[codmun] = str(frente).strip()

                for t in range(12):
                    dcol = f"{MONTHS_ES[t]}_delivered"
                    fcol = f"{MONTHS_ES[t]}_finished"

                    if dcol in df_track.columns:
                        delivered[(codmun, t)] = clamp_int(r.get(dcol, 0), default=0, minv=0)
                    if fcol in df_track.columns:
                        finished[(codmun, t)] = bool(int(r.get(fcol, 0) or 0))

    # 3) Lock plan for finished months (so the editor shows it frozen)
    for codmun in selected_codmun:
        for t in range(12):
            if finished.get((codmun, t), False):
                planned_locked[(codmun, t)] = int(planned_current.get((codmun, t), 0))

    return dict(
        selected_codmun=selected_codmun,
        targets=targets,
        windows=windows,
        delivered=delivered,
        finished=finished,
        planned_locked=planned_locked,
        planned_current=planned_current,
        frontes=frontes,
    )


def dept_spillover_active(
    dept_coddep: int,
    selected_muns: pd.DataFrame,
    finished: Dict[Tuple[int, int], bool],
    delivered: Dict[Tuple[int, int], int],
    targets: Dict[int, int],
    windows: Dict[int, Tuple[int, int]],
) -> bool:
    mun_ids = selected_muns.loc[selected_muns["coddep"] == dept_coddep, "codmun"].astype(int).tolist()
    if not mun_ids:
        return False

    total_remaining = 0
    all_window_finished = True

    for codmun in mun_ids:
        target = targets.get(codmun, 0)
        delivered_sum = sum(delivered.get((codmun, t), 0) for t in range(12))
        rem = max(0, target - delivered_sum)
        total_remaining += rem

        s_idx, e_idx = windows.get(codmun, (0, 11))
        for t in range(s_idx, e_idx + 1):
            if not finished.get((codmun, t), False):
                all_window_finished = False
                break

    return (total_remaining > 0) and all_window_finished


def eligible_months_for_mun(
    codmun: int,
    finished: Dict[Tuple[int, int], bool],
    windows: Dict[int, Tuple[int, int]],
    spillover_active_for_dept: bool,
) -> List[int]:
    s_idx, e_idx = windows.get(codmun, (0, 11))
    if not spillover_active_for_dept:
        return [t for t in range(s_idx, e_idx + 1) if not finished.get((codmun, t), False)]
    else:
        # allow window months + later months in spillover mode
        return [t for t in range(s_idx, 12) if not finished.get((codmun, t), False)]


def build_editor_df(
    selected_muns: pd.DataFrame,
    targets: Dict[int, int],
    windows: Dict[int, Tuple[int, int]],
    delivered: Dict[Tuple[int, int], int],
    finished: Dict[Tuple[int, int], bool],
    planned_locked: Dict[Tuple[int, int], int],
    planned_current: Dict[Tuple[int, int], int],
    frontes: Dict[int, str],
) -> pd.DataFrame:
    rows = []
    for _, r in selected_muns.iterrows():
        codmun = int(r["codmun"])
        row = {
            "frente": frontes.get(codmun, ""),
            "coddep": int(r["coddep"]),
            "departamento": r["departamento"],
            "codmun": codmun,
            "municipio": r["municipio_norm"],
            "target": int(targets.get(codmun, 0)),
        }
        s_idx, e_idx = windows.get(codmun, (0, 11))
        row["start_month"] = IDX_TO_EN[s_idx]
        row["end_month"] = IDX_TO_EN[e_idx]

        for t in range(12):
            row[f"{MONTHS_EN[t]}_delivered"] = int(delivered.get((codmun, t), 0))
            row[f"{MONTHS_EN[t]}_finished"] = bool(finished.get((codmun, t), False))
            if row[f"{MONTHS_EN[t]}_finished"]:
                row[f"{MONTHS_EN[t]}_planned"] = int(planned_locked.get((codmun, t), 0))
            else:
                row[f"{MONTHS_EN[t]}_planned"] = int(planned_current.get((codmun, t), 0))

        rows.append(row)

    df = pd.DataFrame(rows)

    # Sort: frente, coddep, codmun
    df["frente_sort"] = df["frente"].fillna("").astype(str)
    df = df.sort_values(["frente_sort", "coddep", "codmun"]).drop(columns=["frente_sort"])

    base_cols = ["frente", "coddep", "departamento", "codmun", "municipio", "target", "start_month", "end_month"]
    month_cols = []
    for m in MONTHS_EN:
        month_cols += [f"{m}_delivered", f"{m}_finished", f"{m}_planned"]
    return df[base_cols + month_cols]


def sync_state_from_editor_on_apply(
    editor_df: pd.DataFrame,
    targets: Dict[int, int],
    windows: Dict[int, Tuple[int, int]],
    delivered: Dict[Tuple[int, int], int],
    finished: Dict[Tuple[int, int], bool],
    planned_locked: Dict[Tuple[int, int], int],
    planned_current: Dict[Tuple[int, int], int],
    frontes: Dict[int, str],
):
    for _, row in editor_df.iterrows():
        codmun = int(row["codmun"])

        # Frente editable
        frontes[codmun] = str(row.get("frente", "") or "").strip()

        targets[codmun] = clamp_int(row["target"], default=0, minv=0)

        s = row["start_month"]
        e = row["end_month"]
        if s in EN_TO_IDX and e in EN_TO_IDX:
            s_idx = EN_TO_IDX[s]
            e_idx = EN_TO_IDX[e]
            if s_idx <= e_idx:
                windows[codmun] = (s_idx, e_idx)

        for t in range(12):
            dval = clamp_int(row[f"{MONTHS_EN[t]}_delivered"], default=0, minv=0)
            delivered[(codmun, t)] = dval

            fval = bool(row[f"{MONTHS_EN[t]}_finished"])
            prev_f = bool(finished.get((codmun, t), False))
            finished[(codmun, t)] = fval

            if (not prev_f) and fval:
                # lock current planned at time of finish
                planned_locked[(codmun, t)] = int(planned_current.get((codmun, t), 0))


def solve_plan_milp(
    selected_muns: pd.DataFrame,
    targets: Dict[int, int],
    windows: Dict[int, Tuple[int, int]],
    delivered: Dict[Tuple[int, int], int],
    finished: Dict[Tuple[int, int], bool],
    month_coeffs: List[float],
    capacity_soft_ceiling: int,
    deviation_penalty: float = 1.0,
) -> Tuple[Dict[Tuple[int, int], int], Dict]:
    """
    NOTE: Keep your existing MILP behavior; no solver logic changes requested here.
    """
    month_coeffs = np.array(month_coeffs, dtype=float)
    month_coeffs[month_coeffs < 0] = 0.0

    mun_ids = selected_muns["codmun"].astype(int).tolist()
    mun_to_dept = dict(zip(selected_muns["codmun"].astype(int), selected_muns["coddep"].astype(int)))
    dept_ids = sorted(set(selected_muns["coddep"].astype(int).tolist()))

    # remaining
    rem_by_mun = {}
    for codmun in mun_ids:
        target = int(targets.get(codmun, 0))
        delivered_sum = sum(int(delivered.get((codmun, t), 0)) for t in range(12))
        rem_by_mun[codmun] = max(0, target - delivered_sum)

    # dept remaining totals (for diagnostics + weights)
    rem_by_dept = {d: 0 for d in dept_ids}
    for codmun in mun_ids:
        rem_by_dept[mun_to_dept[codmun]] += rem_by_mun[codmun]

    cap = int(capacity_soft_ceiling)

    # spillover per dept (existing rule)
    spillover_by_dept = {}
    for d in dept_ids:
        spillover_by_dept[d] = dept_spillover_active(
            dept_coddep=d,
            selected_muns=selected_muns,
            finished=finished,
            delivered=delivered,
            targets=targets,
            windows=windows,
        )

    # eligible months
    eligible = {}
    for codmun in mun_ids:
        d = mun_to_dept[codmun]
        eligible[codmun] = eligible_months_for_mun(
            codmun=codmun,
            finished=finished,
            windows=windows,
            spillover_active_for_dept=spillover_by_dept[d],
        )

    infeasible_muns = [m for m in mun_ids if rem_by_mun[m] > 0 and len(eligible[m]) == 0]
    if infeasible_muns:
        return {}, {"status": "infeasible", "reason": f"No eligible months for {infeasible_muns[:10]}"}

    # open months per dept
    dept_open_months = {d: set() for d in dept_ids}
    for codmun in mun_ids:
        dept_open_months[mun_to_dept[codmun]].update(eligible[codmun])

    # weights per dept (renormalized domain)
    weights_by_dept = {}
    for d in dept_ids:
        open_months = sorted(dept_open_months[d])
        if not open_months:
            weights_by_dept[d] = np.zeros(12, dtype=float)
            continue

        if spillover_by_dept[d]:
            w = np.zeros(12, dtype=float)
            for t in open_months:
                w[t] = 1.0
        else:
            w = np.zeros(12, dtype=float)
            for t in open_months:
                w[t] = float(month_coeffs[t])

        weights_by_dept[d] = normalize_weights(w)

    desired = {}
    for d in dept_ids:
        w = weights_by_dept[d]
        for t in range(12):
            desired[(d, t)] = float(rem_by_dept[d]) * float(w[t])

    # MILP
    prob = pulp.LpProblem("FloorsPlanningMILP", pulp.LpMinimize)

    x = {}
    for codmun in mun_ids:
        for t in eligible[codmun]:
            x[(codmun, t)] = pulp.LpVariable(f"x_{codmun}_{t}", lowBound=0, cat="Integer")

    S = {(d, t): pulp.LpVariable(f"S_{d}_{t}", lowBound=0, cat="Integer")
         for d in dept_ids for t in range(12)}

    G = {t: pulp.LpVariable(f"G_{t}", lowBound=0, cat="Integer") for t in range(12)}

    dev_pos = {(d, t): pulp.LpVariable(f"devp_{d}_{t}", lowBound=0, cat="Continuous")
               for d in dept_ids for t in range(12)}
    dev_neg = {(d, t): pulp.LpVariable(f"devn_{d}_{t}", lowBound=0, cat="Continuous")
               for d in dept_ids for t in range(12)}

    # municipality totals
    for codmun in mun_ids:
        prob += pulp.lpSum(x[(codmun, t)] for t in eligible[codmun]) == rem_by_mun[codmun], f"mun_total_{codmun}"

    # dept sums
    for d in dept_ids:
        mun_in_dept = [m for m in mun_ids if mun_to_dept[m] == d]
        for t in range(12):
            prob += S[(d, t)] == pulp.lpSum(x.get((m, t), 0) for m in mun_in_dept), f"dept_sum_{d}_{t}"

    # global totals and HARD cap (recommended for consistency with spillover pressure)
    for t in range(12):
        prob += G[t] == pulp.lpSum(x.get((m, t), 0) for m in mun_ids), f"global_sum_{t}"
        prob += G[t] <= cap, f"cap_hard_{t}"

    # deviation
    for d in dept_ids:
        for t in range(12):
            prob += S[(d, t)] - desired[(d, t)] <= dev_pos[(d, t)]
            prob += desired[(d, t)] - S[(d, t)] <= dev_neg[(d, t)]

    prob += deviation_penalty * (pulp.lpSum(dev_pos.values()) + pulp.lpSum(dev_neg.values()))

    status = prob.solve(pulp.PULP_CBC_CMD(msg=False))
    st_name = pulp.LpStatus[status]
    if st_name not in ("Optimal", "Feasible"):
        return {}, {"status": st_name, "reason": "Solver did not find feasible solution"}

    planned_out = {(codmun, t): 0 for codmun in mun_ids for t in range(12)}
    for (codmun, t), var in x.items():
        planned_out[(codmun, t)] = int(round(pulp.value(var) or 0))

    diag = {
        "status": st_name,
        "cap": cap,
        "deviation_total": float(sum((pulp.value(v) or 0) for v in list(dev_pos.values()) + list(dev_neg.values()))),
        "spillover_by_dept": spillover_by_dept,
    }
    return planned_out, diag


# ----------------------------
# Styling (Heatmap by Frente)
# ----------------------------
def build_frente_color_map(frontes: Dict[int, str]) -> Dict[str, str]:
    """
    Map each unique frente to a background color.
    (Fixed palette; deterministic by sorted unique frente.)
    """
    palette = [
        "#E3F2FD", "#E8F5E9", "#FFF3E0", "#F3E5F5", "#E0F7FA",
        "#FCE4EC", "#F1F8E9", "#FFFDE7", "#EDE7F6", "#E0F2F1",
        "#F9FBE7", "#E8EAF6"
    ]
    uniq = sorted({(v or "").strip() for v in frontes.values() if str(v).strip() != ""})
    cmap = {}
    for i, f in enumerate(uniq):
        cmap[f] = palette[i % len(palette)]
    cmap[""] = "#FFFFFF"
    return cmap


def build_plan_heatmap_df(editor_df: pd.DataFrame) -> pd.DataFrame:
    # Make this robust in case editor_df is missing expected columns during a rerun
    needed = ["frente", "coddep", "codmun", "municipio"] + [f"{m}_planned" for m in MONTHS_EN]
    missing = [c for c in needed if c not in editor_df.columns]
    if missing:
        print("missing columns:", missing)
        # fallback: return whatever planned columns exist (no crash)
        planned_cols = [c for c in editor_df.columns if c.endswith("_planned")]
        base_cols = [c for c in ["coddep", "codmun", "municipio"] if c in editor_df.columns]
        return editor_df[base_cols + planned_cols].copy()
    return editor_df[needed].copy()


def plan_heatmap_styler(df: pd.DataFrame, frente_colors: Dict[str, str]):
    planned_cols = [c for c in df.columns if c.endswith("_planned")]

    # Map row index -> frente safely (works even if apply only sees subset columns)
    if "frente" in df.columns:
        frente_by_row = df["frente"].fillna("").astype(str).to_dict()
    else:
        frente_by_row = {}

    def _style_row(row, frente: str):
        frente = str(frente or "").strip()
        base = frente_colors.get(frente, "#FFFFFF")

        styles = []
        for col in row.index:
            if col.endswith("_planned"):
                try:
                    v = int(row[col])
                except Exception:
                    v = 0
                if v > 0:
                    styles.append(f"background-color: {base}; font-weight: 600;")
                else:
                    styles.append("")
            else:
                styles.append("")
        return styles

    # Use lambda to inject frente for this row.name
    sty = df.style.apply(
        lambda row: _style_row(row, frente_by_row.get(row.name, "")),
        axis=1,
        subset=planned_cols
    )
    return sty



# ----------------------------
# Excel update/export (preserve uploaded workbook)
# ----------------------------
def _find_sheet(wb, preferred: str) -> openpyxl.worksheet.worksheet.Worksheet:
    return wb[preferred] if preferred in wb.sheetnames else wb[wb.sheetnames[0]]


def _sheet_headers_map(ws) -> Dict[str, int]:
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        headers[str(v).strip()] = c
    return headers


def export_into_original_workbook_bytes(
    selected_muns: pd.DataFrame,
    targets: Dict[int, int],
    windows: Dict[int, Tuple[int, int]],
    delivered: Dict[Tuple[int, int], int],
    finished: Dict[Tuple[int, int], bool],
    planned_locked: Dict[Tuple[int, int], int],
    planned_current: Dict[Tuple[int, int], int],
    frontes: Dict[int, str],
) -> bytes:
    """
    Loads the workbook the user originally loaded/uploaded and updates Sheet1/Sheet2 in-place.
    Adds totals/accum rows and municipality summary at the end of Sheet1.
    Also keeps/creates a Tracking sheet for delivered/finished.
    """
    if "workbook_bytes" not in st.session_state or not st.session_state.workbook_bytes:
        raise ValueError("No original workbook loaded. Load or upload an Excel first.")

    wb = openpyxl.load_workbook(io.BytesIO(st.session_state.workbook_bytes))

    ws1 = _find_sheet(wb, "Sheet1")
    ws2 = _find_sheet(wb, "Sheet2") if "Sheet2" in wb.sheetnames else None

    # Prepare month header matching in Sheet1/Sheet2
    months_es_headers = MONTHS_ES[:]  # canonical
    # note: Excel may have "Julio " with trailing space; we match by strip
    ws1_headers = _sheet_headers_map(ws1)
    ws2_headers = _sheet_headers_map(ws2) if ws2 else {}

    def find_month_col(headers_map: Dict[str, int], month_es: str) -> int:
        # exact match first
        if month_es in headers_map:
            return headers_map[month_es]
        # strip match
        for k, v in headers_map.items():
            if k.strip() == month_es:
                return v
        return None

    ws1_month_cols = {t: find_month_col(ws1_headers, MONTHS_ES[t]) for t in range(12)}
    ws2_month_cols = {t: find_month_col(ws2_headers, MONTHS_ES[t]) for t in range(12)} if ws2 else {}

    # Build quick lookup for selected rows by Codmun in each sheet
    def build_row_index(ws, codmun_header="Codmun") -> Dict[int, int]:
        headers = _sheet_headers_map(ws)
        if codmun_header not in headers:
            return {}
        col = headers[codmun_header]
        idx = {}
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, col).value
            if v is None:
                continue
            try:
                idx[int(v)] = r
            except Exception:
                continue
        return idx

    ws1_row_by_codmun = build_row_index(ws1, "Codmun")
    ws2_row_by_codmun = build_row_index(ws2, "Codmun") if ws2 else {}

    # Update Sheet1 planned values + Frente + Pisos Proyectados
    for _, r in selected_muns.iterrows():
        codmun = int(r["codmun"])
        if codmun not in ws1_row_by_codmun:
            continue
        row_i = ws1_row_by_codmun[codmun]

        # Update Frente if column exists
        if "Frente" in ws1_headers:
            ws1.cell(row_i, ws1_headers["Frente"]).value = frontes.get(codmun, "")

        # Update target
        if "Pisos Proyectados" in ws1_headers:
            ws1.cell(row_i, ws1_headers["Pisos Proyectados"]).value = int(targets.get(codmun, 0))

        # Update month planned numbers
        for t in range(12):
            c = ws1_month_cols.get(t)
            if c is None:
                continue
            if finished.get((codmun, t), False):
                p = int(planned_locked.get((codmun, t), 0))
            else:
                p = int(planned_current.get((codmun, t), 0))
            ws1.cell(row_i, c).value = p

    # Update Sheet2: Frente, Pisos Proyectados, window flags, Meses, Pisos/Mes
    if ws2 is not None and ws2_headers:
        for _, r in selected_muns.iterrows():
            codmun = int(r["codmun"])
            if codmun not in ws2_row_by_codmun:
                continue
            row_i = ws2_row_by_codmun[codmun]

            if "Frente" in ws2_headers:
                ws2.cell(row_i, ws2_headers["Frente"]).value = frontes.get(codmun, "")
            if "Pisos Proyectados" in ws2_headers:
                ws2.cell(row_i, ws2_headers["Pisos Proyectados"]).value = int(targets.get(codmun, 0))

            s_idx, e_idx = windows.get(codmun, (0, 11))
            # month flags (1 inside window, blank outside)
            for t in range(12):
                c = ws2_month_cols.get(t)
                if c is None:
                    continue
                ws2.cell(row_i, c).value = 1 if (s_idx <= t <= e_idx) else None

            # Meses and Pisos/Mes
            if "Meses" in ws2_headers:
                ws2.cell(row_i, ws2_headers["Meses"]).value = int(e_idx - s_idx + 1)
            if "Pisos/Mes" in ws2_headers:
                months = max(1, int(e_idx - s_idx + 1))
                ws2.cell(row_i, ws2_headers["Pisos/Mes"]).value = int(round(int(targets.get(codmun, 0)) / months))

    # Add/Update Tracking sheet for delivered/finished
    if "Tracking" in wb.sheetnames:
        wst = wb["Tracking"]
        wst.delete_rows(1, wst.max_row)
    else:
        wst = wb.create_sheet("Tracking")

    # Write Tracking header
    tracking_headers = ["Frente", "Coddep", "Departamento", "Codmun", "Municipio", "Pisos Proyectados"]
    for t in range(12):
        tracking_headers += [f"{MONTHS_ES[t]}_delivered", f"{MONTHS_ES[t]}_finished"]
    for j, h in enumerate(tracking_headers, start=1):
        wst.cell(1, j).value = h
        wst.cell(1, j).font = Font(bold=True)

    # Tracking rows
    row = 2
    for _, r in selected_muns.sort_values(["coddep", "codmun"]).iterrows():
        codmun = int(r["codmun"])
        base = [
            frontes.get(codmun, ""),
            int(r["coddep"]),
            r["departamento"],
            codmun,
            r["municipio_norm"],
            int(targets.get(codmun, 0)),
        ]
        vals = base[:]
        for t in range(12):
            vals += [int(delivered.get((codmun, t), 0)), int(bool(finished.get((codmun, t), False)))]
        for j, v in enumerate(vals, start=1):
            wst.cell(row, j).value = v
        row += 1

    # Append totals + accum + municipality summary to bottom of Sheet1
    # Find last data row by scanning Codmun column.
    codmun_col = ws1_headers.get("Codmun")
    last_data_row = ws1.max_row
    if codmun_col:
        last_data_row = 1
        for r in range(2, ws1.max_row + 1):
            v = ws1.cell(r, codmun_col).value
            if v is not None:
                last_data_row = r

    start_row = last_data_row + 2  # blank row separator

    def write_row(label: str, values_by_month: List[int], row_i: int):
        # label in first column
        ws1.cell(row_i, 1).value = label
        ws1.cell(row_i, 1).font = Font(bold=True)
        for t in range(12):
            c = ws1_month_cols.get(t)
            if c is not None:
                ws1.cell(row_i, c).value = int(values_by_month[t])

    # Compute totals from current session_state (not Excel formulas)
    planned_tot = [0] * 12
    deliv_tot = [0] * 12
    for _, r in selected_muns.iterrows():
        codmun = int(r["codmun"])
        for t in range(12):
            deliv_tot[t] += int(delivered.get((codmun, t), 0))
            if finished.get((codmun, t), False):
                planned_tot[t] += int(planned_locked.get((codmun, t), 0))
            else:
                planned_tot[t] += int(planned_current.get((codmun, t), 0))

    plan_acc = np.cumsum(planned_tot).astype(int).tolist()
    deliv_acc = np.cumsum(deliv_tot).astype(int).tolist()
    diff_tot = (np.array(planned_tot) - np.array(deliv_tot)).astype(int).tolist()

    write_row("TOTAL_PLAN", planned_tot, start_row)
    write_row("ACCUM_PLAN", plan_acc, start_row + 1)
    write_row("TOTAL_DELIV", deliv_tot, start_row + 2)
    write_row("ACCUM_DELIV", deliv_acc, start_row + 3)
    write_row("DIFF_PLAN_MINUS_DELIV", diff_tot, start_row + 4)

    # Municipality summary (count + list)
    mun_list = selected_muns.sort_values(["coddep", "codmun"])["municipio_norm"].astype(str).tolist()
    ws1.cell(start_row + 6, 1).value = "MUNICIPALITIES_COUNT"
    ws1.cell(start_row + 6, 1).font = Font(bold=True)
    ws1.cell(start_row + 6, 2).value = int(len(mun_list))

    ws1.cell(start_row + 7, 1).value = "MUNICIPALITIES_LIST"
    ws1.cell(start_row + 7, 1).font = Font(bold=True)
    ws1.cell(start_row + 7, 2).value = ", ".join(mun_list)

    # Optional: color planned cells in Sheet1 by Frente (for easier scan)
    frente_colors = build_frente_color_map(frontes)
    # Convert hex to openpyxl fill
    def fill_from_hex(hex_color: str) -> PatternFill:
        hc = hex_color.lstrip("#")
        # openpyxl expects ARGB; use FF alpha
        return PatternFill("solid", fgColor=f"FF{hc.upper()}")

    if "Frente" in ws1_headers:
        frente_col = ws1_headers["Frente"]
        for _, r in selected_muns.iterrows():
            codmun = int(r["codmun"])
            rr = ws1_row_by_codmun.get(codmun)
            if not rr:
                continue
            f = str(ws1.cell(rr, frente_col).value or "").strip()
            fill = fill_from_hex(frente_colors.get(f, "#FFFFFF"))
            for t in range(12):
                c = ws1_month_cols.get(t)
                if c is None:
                    continue
                v = ws1.cell(rr, c).value
                try:
                    vv = int(v or 0)
                except Exception:
                    vv = 0
                if vv > 0:
                    ws1.cell(rr, c).fill = fill

    # call it
    upsert_config_sheet(
        wb,
        month_coeffs=st.session_state.month_coeffs,
        cap=int(st.session_state.capacity),
    )

    # Save to bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ----------------------------
# Streamlit App
# ----------------------------
st.set_page_config(page_title="Planificador Pisos", layout="wide")
st.title("Planificador Pisos")

CATALOG_PATH = "municipalidades_listado.json"
EXCEL_DEFAULT_PATH = "mds_mano_a_mano_planificacion.xlsx"

mun_df = load_municipios(CATALOG_PATH)

def _init_state():
    st.session_state.selected_codmun = set()
    st.session_state.targets = {}
    st.session_state.windows = {}
    st.session_state.delivered = {}
    st.session_state.finished = {}
    st.session_state.planned_locked = {}
    st.session_state.planned_current = {}
    st.session_state.frontes = {}
    st.session_state.month_coeffs = [10, 10, 12, 12, 14, 14, 16, 16, 12, 10, 8, 6]
    st.session_state.capacity = 20000
    st.session_state.last_diag = None
    st.session_state.workbook_bytes = None
    st.session_state.workbook_filename = "planning.xlsx"

if "selected_codmun" not in st.session_state:
    _init_state()


# ----------------------------
# 1) Load / Setup
# ----------------------------
st.subheader("1) Configuración")

c1, c2, c3 = st.columns([1.2, 1.4, 2.0])

with c1:
    if st.button("Cargar Excel base"):
        with open(EXCEL_DEFAULT_PATH, "rb") as f:
            b = f.read()
        loaded = load_state_from_excel_bytes(b, EXCEL_DEFAULT_PATH, mun_df)
        for k, v in loaded.items():
            st.session_state[k] = v
        st.success("Cargado desde Excel base.")

with c2:
    upl = st.file_uploader("Ó subir un Excel", type=["xlsx"])
    if upl is not None and st.button("Subir Excel"):
        b = upl.read()
        loaded = load_state_from_excel_bytes(b, getattr(upl, "name", "uploaded.xlsx"), mun_df)
        for k, v in loaded.items():
            st.session_state[k] = v
        st.success("Cargado desde Excel subido.")

with c3:
    st.caption("Capacidad global: valor máximo de entregas por mes.")
    st.session_state.capacity = st.number_input(
        "Cupones por mes",
        min_value=0,
        value=int(st.session_state.capacity),
        step=1000
    )

st.markdown("---")


# ----------------------------
# 2) Add municipalities (optional)
# ----------------------------
st.subheader("2) Agregar municipios")

colA, colB, colC = st.columns([1.2, 2.2, 1.2])
with colA:
    dept = st.selectbox("Departmento", sorted(mun_df["departamento"].unique()))
with colB:
    dept_muns = mun_df[mun_df["departamento"] == dept].sort_values("municipio_norm")
    choices = st.multiselect(
        "Municipios",
        options=dept_muns["codmun"].astype(int).tolist(),
        format_func=lambda codmun: f"{dept_muns.loc[dept_muns['codmun']==codmun,'municipio_norm'].iloc[0]}"
    )
with colC:
    if st.button("Agregar seleccionado"):
        for codmun in choices:
            codmun = int(codmun)
            st.session_state.selected_codmun.add(codmun)
            st.session_state.targets.setdefault(codmun, 0)
            st.session_state.windows.setdefault(codmun, (0, 11))
            st.session_state.frontes.setdefault(codmun, "")
            for t in range(12):
                st.session_state.delivered.setdefault((codmun, t), 0)
                st.session_state.finished.setdefault((codmun, t), False)
                st.session_state.planned_current.setdefault((codmun, t), 0)
        st.success("Agregado.")

if st.button("Eliminar todos los municipios / Restaurar"):
    _init_state()
    st.success("Restaurado.")

st.markdown("---")


# ----------------------------
# 3) Efficiency coefficients
# ----------------------------
st.subheader("3) Coeficientes de eficiencia")
coeff_cols = st.columns(6)
new_coeffs = []
for i in range(12):
    with coeff_cols[i % 6]:
        new_coeffs.append(
            st.number_input(
                f"{MONTHS_EN[i]}",
                min_value=0.0,
                value=float(st.session_state.month_coeffs[i]),
                step=1.0
            )
        )
st.session_state.month_coeffs = new_coeffs

st.markdown("---")


# ----------------------------
# 4) Main editor (two-stage form)
# ----------------------------
selected_codmun = sorted(list(st.session_state.selected_codmun))
selected_muns_df = mun_df[mun_df["codmun"].isin(selected_codmun)].copy()

st.subheader("4) Tabla de planificación mensual")

if selected_muns_df.empty:
    st.info("Carga desde un Excel o inicia agregando un municipio.")
    st.stop()

editor_df = build_editor_df(
    selected_muns=selected_muns_df,
    targets=st.session_state.targets,
    windows=st.session_state.windows,
    delivered=st.session_state.delivered,
    finished=st.session_state.finished,
    planned_locked=st.session_state.planned_locked,
    planned_current=st.session_state.planned_current,
    frontes=st.session_state.frontes,
)

st.markdown("**Marca el mes completo para los municipios (terminado/no terminado):**")
finish_cols = st.columns(12)

toggle_clicked = None
toggle_to_finished = None  # True = finish all, False = unfinish all

for t, m in enumerate(MONTHS_EN):
    # if every municipality is finished this month, we show Unfinish
    all_finished = all(
        bool(st.session_state.finished.get((codmun, t), False))
        for codmun in selected_codmun
    )

    label = f"↩️ {m}" if all_finished else f"✅ {m}"

    with finish_cols[t]:
        if st.button(label, key=f"toggle_finish_all_{m}"):
            toggle_clicked = t
            toggle_to_finished = (not all_finished)

if toggle_clicked is not None:
    t = toggle_clicked

    if toggle_to_finished:
        # FINISH ALL: mark finished + lock plan snapshot
        for codmun in selected_codmun:
            st.session_state.finished[(codmun, t)] = True
            st.session_state.planned_locked[(codmun, t)] = int(
                st.session_state.planned_current.get((codmun, t), 0)
            )
        st.success(f"{MONTHS_EN[t]} marcado como terminado para todos los municipios.")
    else:
        # UNFINISH ALL: mark unfinished + unlock (remove lock snapshot)
        for codmun in selected_codmun:
            st.session_state.finished[(codmun, t)] = False
            st.session_state.planned_locked.pop((codmun, t), None)
        st.success(f"{MONTHS_EN[t]} marcado como no terminado para todos los municipios.")

    st.rerun()



with st.form("top_table_form", clear_on_submit=False):
    st.caption("Ajusta → luego presiona **Aplicar cambios**. Columns planificada no son modificables")
    edited = st.data_editor(
        editor_df,
        use_container_width=True,
        num_rows="fixed",
        hide_index=True,
        column_config={
            "start_month": st.column_config.SelectboxColumn(options=MONTHS_EN),
            "end_month": st.column_config.SelectboxColumn(options=MONTHS_EN),
            **{f"{m}_planned": st.column_config.NumberColumn(disabled=True) for m in MONTHS_EN},
        }
    )
    btnA, btnB, btnC = st.columns([1.2, 1.6, 2.2])
    apply_clicked = btnA.form_submit_button("Aplicar cambios")
    recalc_clicked = btnB.form_submit_button("Recalcular plan")
    export_clicked = btnC.form_submit_button("Preparar Excel para descarga")

if apply_clicked:
    sync_state_from_editor_on_apply(
        editor_df=edited,
        targets=st.session_state.targets,
        windows=st.session_state.windows,
        delivered=st.session_state.delivered,
        finished=st.session_state.finished,
        planned_locked=st.session_state.planned_locked,
        planned_current=st.session_state.planned_current,
        frontes=st.session_state.frontes,
    )
    st.success("Cambios aplicados.")

if recalc_clicked:
    sync_state_from_editor_on_apply(
        editor_df=edited,
        targets=st.session_state.targets,
        windows=st.session_state.windows,
        delivered=st.session_state.delivered,
        finished=st.session_state.finished,
        planned_locked=st.session_state.planned_locked,
        planned_current=st.session_state.planned_current,
        frontes=st.session_state.frontes,
    )

    planned_out, diag = solve_plan_milp(
        selected_muns=selected_muns_df,
        targets=st.session_state.targets,
        windows=st.session_state.windows,
        delivered=st.session_state.delivered,
        finished=st.session_state.finished,
        month_coeffs=st.session_state.month_coeffs,
        capacity_soft_ceiling=int(st.session_state.capacity),
    )
    st.session_state.last_diag = diag

    if diag.get("status") == "infeasible":
        st.error(diag.get("reason", "Imposible"))
    elif planned_out:
        for (codmun, t), v in planned_out.items():
            if not st.session_state.finished.get((codmun, t), False):
                st.session_state.planned_current[(codmun, t)] = int(v)
        st.success(f"Resuleto: {diag.get('status')}")
        st.rerun()
    else:
        st.error(f"Error al calcular: {diag.get('status')} | {diag.get('reason')}")

export_bytes = None
if export_clicked:
    sync_state_from_editor_on_apply(
        editor_df=edited,
        targets=st.session_state.targets,
        windows=st.session_state.windows,
        delivered=st.session_state.delivered,
        finished=st.session_state.finished,
        planned_locked=st.session_state.planned_locked,
        planned_current=st.session_state.planned_current,
        frontes=st.session_state.frontes,
    )

    export_bytes = export_into_original_workbook_bytes(
        selected_muns=selected_muns_df,
        targets=st.session_state.targets,
        windows=st.session_state.windows,
        delivered=st.session_state.delivered,
        finished=st.session_state.finished,
        planned_locked=st.session_state.planned_locked,
        planned_current=st.session_state.planned_current,
        frontes=st.session_state.frontes,
    )
    st.success("Excel a exportar listo (descarga abajo).")

st.markdown("---")


# ----------------------------
# 5) Summaries
# ----------------------------
st.subheader("5) Resumenes")

col1, col2 = st.columns(2)

with col1:
    st.markdown("**Restantes por municipio**")
    rem_rows = []
    # include frente and keep sorted
    tmp = selected_muns_df.copy()
    tmp["frente"] = tmp["codmun"].map(lambda c: st.session_state.frontes.get(int(c), ""))
    tmp = tmp.sort_values(["frente", "coddep", "codmun"])

    for _, r in tmp.iterrows():
        codmun = int(r["codmun"])
        target = int(st.session_state.targets.get(codmun, 0))
        deliv = sum(int(st.session_state.delivered.get((codmun, t), 0)) for t in range(12))
        rem = max(0, target - deliv)
        rem_rows.append({
            "frente": st.session_state.frontes.get(codmun, ""),
            "coddep": int(r["coddep"]),
            "departamento": r["departamento"],
            "codmun": codmun,
            "municipio": r["municipio_norm"],
            "target": target,
            "delivered_total": deliv,
            "remaining": rem
        })
    st.dataframe(pd.DataFrame(rem_rows), use_container_width=True, hide_index=True)

with col2:
    st.markdown("**Meses totales (planeados vs entregados + acumulados)**")
    month_rows = []
    cum_plan = 0
    cum_delv = 0

    for t in range(12):
        planned_total = 0
        delivered_total = 0

        for codmun in selected_codmun:
            delivered_total += int(st.session_state.delivered.get((codmun, t), 0))
            if st.session_state.finished.get((codmun, t), False):
                planned_total += int(st.session_state.planned_locked.get((codmun, t), 0))
            else:
                planned_total += int(st.session_state.planned_current.get((codmun, t), 0))

        diff = planned_total - delivered_total
        cum_plan += planned_total
        cum_delv += delivered_total

        month_rows.append({
            "month": MONTHS_EN[t],
            "planned_total": planned_total,
            "delivered_total": delivered_total,
            "diff": diff,
            "planned_cum": cum_plan,
            "delivered_cum": cum_delv,
            "diff_cum": cum_plan - cum_delv
        })

    st.dataframe(pd.DataFrame(month_rows), use_container_width=True, hide_index=True)

st.subheader("6) Mapa de calor de mapa (pintados por frente)")
frente_colors = build_frente_color_map(st.session_state.frontes)

heat_df = build_plan_heatmap_df(editor_df)
st.dataframe(plan_heatmap_styler(heat_df, frente_colors), use_container_width=True, hide_index=True)

if export_bytes is not None:
    filename = st.session_state.workbook_filename or "planning.xlsx"
    st.download_button(
        label="Descarga Excel actualizado",
        data=export_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
